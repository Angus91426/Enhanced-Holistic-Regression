"""
Simulation.py
=============

Functions
---------
run_detection_simulation : synthetic multicollinearity-detection ablation study
                            (Methods A-E, SMR, Bertsimas) plus the paired
                            significance tests on the produced Performance table.
run_reduction_simulation : dimensionality-reduction screen comparison
                            (Eigvec vs Corr at several outlier thresholds).
run_realworld_detection  : detection on pre-processed real-world datasets
                            (SMR and/or Bertsimas).
"""

import os, time, math, datetime

import numpy as np, pandas as pd
import Scalable_Multicollinearity_Recovery as SMR

from scipy.stats import wilcoxon, ttest_rel
from rich import print
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import warnings
warnings.filterwarnings( 'ignore' )

# MR groups (and their column pairs) always appear in this order.
MR_ORDER = ['MR2', 'MR3', 'MR4', 'MR8', 'MR10', 'MR15', 'MR20']

# One row per detection event in the per-simulation detected sheet.
DETECTED_HEADERS = ['Method', 'Status', 'Detected_Relationship', 'Norm', 'Coefficient', 'Gap', 'Time']

# Metric columns for the significance test and whether higher or lower is better.
METRIC_DIRECTION = {
    'MACC': 'higher_is_better',
    'MFPR': 'lower_is_better',
    'Time': 'lower_is_better',
}


# --------------------------------------------------------------------------- #
# Shared helpers
# --------------------------------------------------------------------------- #
def _write_log_banner( log_file: str, message: str ) -> None:
    """Append a timestamped banner line to the live Gurobi log file."""
    if log_file is None:
        return
    with open( log_file, 'a', buffering = 1 ) as f:
        f.write( '\n' + '=' * 80 + '\n' )
        f.write( f'[{datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}] {message}\n' )
        f.write( '=' * 80 + '\n' )


def _support_to_feature_str( support, feature_col: np.ndarray ) -> str:
    """Convert feature indices to a comma-separated feature-name string."""
    if support is None or len( support ) == 0:
        return ''
    return ', '.join( str( feature_col[ int( i ) ] ) for i in support )


def _coef_to_str( A_opt ) -> str:
    """Convert an A_opt eigenvector to a 6-decimal comma-separated string."""
    if A_opt is None:
        return None
    return ', '.join( f'{float( a ):.6f}' for a in A_opt )


def _build_detected_sheet( workbook: Workbook, sheet_name: str, rows: list ) -> None:
    """Append a sheet listing every detection event of one simulation."""
    ws = workbook.create_sheet( title = sheet_name )
    for col_idx, header in enumerate( DETECTED_HEADERS, start = 1 ):
        ws.cell( row = 1, column = col_idx, value = header )
    for row_offset, row_data in enumerate( rows, start = 2 ):
        for col_idx, header in enumerate( DETECTED_HEADERS, start = 1 ):
            ws.cell( row = row_offset, column = col_idx, value = row_data.get( header ) )


def _build_simulation_sheet( workbook: Workbook, sheet_name: str, MR: list,
                            original_col: dict, coef_dict: dict,
                            feature_col: np.ndarray ) -> None:
    """Append a sheet describing every true relationship of one simulation.
    
    Each relationship occupies one (Feature, Coefficient) column pair, grouped
    by MR type. The LHS feature (X[:, idx[0]] = X[:, idx[1:]] @ coef[1:]) gets
    coefficient -1.
    """
    ws = workbook.create_sheet( title = sheet_name )
    col_cursor = 1
    for mr_name, mr_count in zip( MR_ORDER, MR ):
        if mr_count == 0:
            continue
        relationships_idx = original_col.get( mr_name, [] )
        relationships_coef = coef_dict.get( mr_name, [] )
        for rel_idx, rel_coef in zip( relationships_idx, relationships_coef ):
            feat_col_letter = get_column_letter( col_cursor )
            coef_col_letter = get_column_letter( col_cursor + 1 )
            
            # Row 1: relationship-type label; Row 2: sub-headers.
            ws[f'{feat_col_letter}1'] = mr_name
            ws[f'{coef_col_letter}1'] = mr_name
            ws[f'{feat_col_letter}2'] = 'Feature'
            ws[f'{coef_col_letter}2'] = 'Coefficient'
            
            # Rows 3+: feature labels and gamma coefficients (LHS coef = -1).
            for row_offset, ( feat_index, gamma ) in enumerate( zip( rel_idx, [-1] + list( rel_coef[1:] ) ) ):
                ws.cell( row = 3 + row_offset, column = col_cursor,     value = feature_col[ feat_index ] )
                ws.cell( row = 3 + row_offset, column = col_cursor + 1, value = float( gamma ) )
            col_cursor += 2


def _stars( p ) -> str:
    """Significance stars for a p-value."""
    if p is None or ( isinstance( p, float ) and np.isnan( p ) ):
        return ''
    if p < 0.001: return '***'
    if p < 0.01:  return '**'
    if p < 0.05:  return '*'
    return 'ns'


def _paired_test( a, b ) -> dict:
    """Paired descriptive stats and p-values for treatment a vs baseline b."""
    a = np.asarray( a, dtype = float )
    b = np.asarray( b, dtype = float )
    diff = a - b
    out = {
        'n': len( a ),
        'mean_treat': a.mean(), 'std_treat': a.std( ddof = 1 ) if len( a ) > 1 else 0.0,
        'median_treat': float( np.median( a ) ),
        'mean_base': b.mean(), 'std_base': b.std( ddof = 1 ) if len( b ) > 1 else 0.0,
        'median_base': float( np.median( b ) ),
        'mean_diff': diff.mean(),
        'wilcoxon_p': np.nan, 'ttest_p': np.nan,
    }
    if np.allclose( diff, 0 ):        # identical samples -> nothing to test
        return out
    try:
        _, out['wilcoxon_p'] = wilcoxon( a, b )
    except ValueError:
        pass
    try:
        _, out['ttest_p'] = ttest_rel( a, b )
    except Exception:
        pass
    return out


def _run_significance_test( df: pd.DataFrame, output_dir: str, baseline: str, methods: list ) -> None:
    """Paired Wilcoxon / t-tests of each method against ``baseline``.
    
    Only coef_min == 0 is tested: the ``baseline`` (Original) is run only at
    that level, so it is the only setting with paired samples. Methods are
    paired by Simulation_id; results go to ``significance_results.csv``.
    """
    present_methods = [ m for m in methods if m in set( df['Method'].unique() ) ]
    if baseline not in present_methods:
        print( f'[significance] baseline {baseline!r} absent; skipping test.' )
        return
    
    rows = []
    for coef in [ 0.0 ]:
        sub = df[df['Coef_min'] == coef]
        base = sub[sub['Method'] == baseline].set_index( 'Simulation_id' )
        for method in present_methods:
            if method == baseline:
                continue
            treat = sub[sub['Method'] == method].set_index( 'Simulation_id' )
            common = sorted( set( base.index ) & set( treat.index ) )   # pair by sim id
            if len( common ) < 2:
                continue
            for metric in METRIC_DIRECTION:
                if metric not in base.columns or metric not in treat.columns:
                    continue
                a = treat.loc[common, metric].to_numpy()
                b = base.loc[common, metric].to_numpy()
                r = _paired_test( a, b )
                rows.append( {
                    'coef_min': coef, 'method': method, 'baseline': baseline,
                    'metric': metric, 'direction': METRIC_DIRECTION.get( metric, '' ),
                    'n': r['n'],
                    'mean_treat': r['mean_treat'], 'std_treat': r['std_treat'], 'median_treat': r['median_treat'],
                    'mean_base': r['mean_base'], 'std_base': r['std_base'], 'median_base': r['median_base'],
                    'mean_diff': r['mean_diff'],
                    'wilcoxon_p': r['wilcoxon_p'], 'wilcoxon_sig': _stars( r['wilcoxon_p'] ),
                    'ttest_p': r['ttest_p'], 'ttest_sig': _stars( r['ttest_p'] ),
                } )
    
    if rows:
        out_path = os.path.join( output_dir, 'significance_results.csv' )
        pd.DataFrame( rows ).to_csv( out_path, index = False, encoding = 'utf-8-sig' )
        print( f'Saved {out_path}' )
    else:
        print( '[significance] no paired comparisons produced.' )


# --------------------------------------------------------------------------- #
# 1. Synthetic detection ablation study + significance test
# --------------------------------------------------------------------------- #
def run_detection_simulation(
    BASE_DIR: str,
    n: int, p: int, MR: list,
    noise_scale: float = 0.01,
    coef_mins = ( 0.0, ),
    simulations: int = 10,
    seed: int = 911122,
    outlier_threshold: int = 5,
    norm_threshold: float = 10 ** ( -2 ),
    alpha: float = 0.5,
    total_time_limit: int = None,
    per_solve_time_limit: int = None,
    run_A: bool = True, run_B: bool = True, run_C: bool = True, run_D: bool = True, run_E: bool = True,
    run_SMR_Corr: bool = True, run_SMR_Eigen: bool = True, run_Bertsimas: bool = True,
    save_relationship: bool = True,
    save_trace: bool = False,
    run_significance: bool = True,
    significance_baseline: str = 'Original',
    csv_id: str = None,
):
    """Run the multicollinearity-detection ablation study, then paired tests.
    
    Parameters
    ----------
    BASE_DIR : str
        Root directory for all outputs (created if missing).
    n, p : int
        Design-matrix sample size and feature count.
    MR : list
        Relationship counts [MR2, MR3, MR4, MR8, MR10, MR15, MR20].
    noise_scale : float
        Std. of the additive Gaussian noise in Simulation_Data.
    coef_mins : iterable of float
        Minimum-|coefficient| levels; one workbook is produced per level.
    simulations : int
        Number of random datasets per (coef_min) setting.
    seed : int
        Seed for the per-setting data-seed generator (reproducibility).
    outlier_threshold : int
        z-score threshold for the correlation screen (Methods A / E / En-Corr).
    norm_threshold, alpha : float
        Inequality norm threshold and eigenvalue-cutoff scale for detection.
    total_time_limit, per_solve_time_limit : int or None
        Overall and per-solve Gurobi time budgets (seconds).
    run_* : bool
        Enable/disable each method.
    save_relationship : bool
        Also write the ground-truth and detected-relationship workbooks.
    save_trace : bool
        Write a separate trace CSV per method and simulation (in addition to the
        consolidated detected workbook). Off by default.
    run_significance : bool
        Run paired Wilcoxon / t-tests on the produced Performance table.
    significance_baseline : str
        Method every other method is compared against.
    csv_id : str or None
        Run tag for the output folder; a timestamp is generated if None.
    """
    csv_id = csv_id or time.strftime( '%Y%m%d-%H%M%S', time.localtime() )
    
    OUTPUT_DIR = os.path.join( BASE_DIR, 'Detection', f'p_{p}_{csv_id}' )
    os.makedirs( OUTPUT_DIR, exist_ok = True )
    log_path = 'gurobi_live.log'
    
    rng = np.random.default_rng( seed = seed )
    
    # ----------------------------------------------------------------------------------------------------------------- #
    # Ablation methods. Each is one fixed Multicollinear config; adjacent comparisons differ by exactly one subprocess. #
    #   A              : Correlation screen only                                                                        #
    #   B              : Eigenvector screen only                                                                        #
    #   C              : Inequality inspection only                                                                     #
    #   D              : Inequality + Irreducibility                                                                    #
    #   E              : Corr screen + Inequality + Irreducibility                                                      #
    #   Enhanced Corr  : Corr screen + Inequality + Irreducibility + Fast-path                                          #
    #   Enhanced Eigen : Eig screen + Inequality + Irreducibility + Fast-path                                           #
    # ----------------------------------------------------------------------------------------------------------------- #
    METHODS = [
        { 'label': 'Method A', 'enabled': run_A, 'flags': {
            'reduction': True,  'reduction_method': 'corr', 'outlier_threshold': outlier_threshold,
            'Inequality_Inspection': False, 'Irreducibility_Inspection': False, 'fastpath': False } },
        
        { 'label': 'Method B', 'enabled': run_B, 'flags': {
            'reduction': True,  'reduction_method': 'eigvec',
            'Inequality_Inspection': False, 'Irreducibility_Inspection': False, 'fastpath': False } },
        
        { 'label': 'Method C', 'enabled': run_C, 'flags': {
            'reduction': False, 'reduction_method': 'eigvec',
            'Inequality_Inspection': True,  'Irreducibility_Inspection': False, 'fastpath': False } },
        
        { 'label': 'Method D', 'enabled': run_D, 'flags': {
            'reduction': False, 'reduction_method': 'eigvec',
            'Inequality_Inspection': True,  'Irreducibility_Inspection': True,  'fastpath': False } },
        
        { 'label': 'Method E', 'enabled': run_E, 'flags': {
            'reduction': True,  'reduction_method': 'corr', 'outlier_threshold': outlier_threshold,
            'Inequality_Inspection': True,  'Irreducibility_Inspection': True,  'fastpath': False } },
        
        { 'label': 'SMR-Corr', 'enabled': run_SMR_Corr, 'flags': {
            'reduction': True,  'reduction_method': 'corr', 'outlier_threshold': outlier_threshold,
            'Inequality_Inspection': True,  'Irreducibility_Inspection': True,  'fastpath': True } },
        
        { 'label': 'SMR-Eigen', 'enabled': run_SMR_Eigen, 'flags': {
            'reduction': True,  'reduction_method': 'eigvec',
            'Inequality_Inspection': True,  'Irreducibility_Inspection': True,  'fastpath': True } },
    ]
    
    # Truncate the live log file so we don't accumulate across runs.
    open( log_path, 'w' ).close()
    _write_log_banner( log_path, f'Simulation start | n={n} p={p} MR={MR} coef_mins={list( coef_mins )} simulations={simulations}' )
    
    performance_ = { 'Coef_min': [], 'Simulation_id': [], 'Method': [], 'MACC': [], 'MFPR': [], 'Time': [] }
    
    def _record_performance( coef_min, data_id, label, ACC, FPR, elapsed ):
        performance_['Coef_min'].append( coef_min )
        performance_['Simulation_id'].append( data_id + 1 )
        performance_['Method'].append( label )
        performance_['MACC'].append( ACC )
        performance_['MFPR'].append( FPR )
        performance_['Time'].append( round( elapsed, 2 ) )
    
    for coef_min in coef_mins:
        data_seeds = rng.integers( low = 0, high = 10000, size = simulations )
        
        # One workbook per coef_min; sheets accumulate over the simulations loop.
        wb = Workbook(); wb.remove( wb.active )
        detected_wb = Workbook(); detected_wb.remove( detected_wb.active )
        
        for data_id, data_seed in enumerate( data_seeds ):
            print( '=.' * 50 )
            print( f'Coef_min: {coef_min}, Data: {data_id + 1} / {simulations}' )
            
            # Generate simulation dataset.
            original_col, coef_dict, X, feature_col = SMR.Simulation_Data( n, p, MR, noise_scale, data_seed, coef_min = coef_min )
            
            # Show the true collinearity relationships.
            true_feat = []
            for key, value in original_col.items():
                print( f'{key}: {value}' )
                for feat_ in value:
                    true_feat += feat_
            print( f'Total number of features involved in multicollinearity: {len( true_feat )}' )
            
            if save_relationship:
                _build_simulation_sheet( workbook = wb, sheet_name = f'Sim_{data_id + 1}', MR = MR,
                                            original_col = original_col, coef_dict = coef_dict, feature_col = feature_col )
            
            # Detection events for THIS simulation, all methods stacked.
            detected_rows = []
            
            # ---- Ablation methods ----
            for method in METHODS:
                if not method['enabled']:
                    continue
                label, flags = method['label'], method['flags']
                print( '{:-^100}'.format( label ) )
                log_tag = f'sim {data_id + 1}/{simulations} | coef_min={coef_min} | {label}'
                trace_path = os.path.join( OUTPUT_DIR, f'Coef_min_{coef_min}_Sim_{data_id + 1}_{label}_Trace.csv' ) if save_trace else None
                
                # Fresh trace=[] per instantiation so each method records only its own events.
                multi_ = SMR.Multicollinear( **flags, norm_threshold = norm_threshold, alpha = alpha,
                                                total_time_limit = total_time_limit, per_solve_time_limit = per_solve_time_limit,
                                                log_file = log_path, log_tag = log_tag, trace = [], trace_file = trace_path )
                start = time.time()
                z_pos = multi_.Ablation_Detection( X = X, col_names = feature_col )
                end = time.time()
                
                ACC, FPR = multi_.Multicollinear_score( z_pos, original_col )
                print( f'ACC: {round( ACC, 2 )}, MFPR: {round( FPR, 2 )}, execution time: {round( end - start, 2 )}' )
                _record_performance( coef_min, data_id, label, ACC, FPR, end - start )
                
                # Norm / Coefficient are blank for methods without inequality inspection.
                # Gap / Time are the MIPGap and solver time of each Minimum_Support solve.
                show_norm = flags['Inequality_Inspection']
                for entry in multi_.trace:
                    detected_rows.append( {
                        'Method': label,
                        'Status': entry['status'],
                        'Detected_Relationship': _support_to_feature_str( entry['support'], feature_col ),
                        'Norm': entry['norm'] if show_norm else None,
                        'Coefficient': _coef_to_str( entry['A_opt'] ) if show_norm else None,
                        'Gap': entry['gap'],
                        'Time': entry['time'],
                    } )
            
            # ---- Bertsimas / Original baseline (big-M formulation) ----
            # Original is only run at coef_min == 0 (the baseline setting).
            if run_Bertsimas and coef_min == 0.0:
                print( '{:-^100}'.format( 'Original' ) )
                log_tag = f'sim {data_id + 1}/{simulations} | coef_min={coef_min} | Original'
                # trace_file so Bertsimas_Minimum_Support records its MIPGap / solver time too.
                trace_path = os.path.join( OUTPUT_DIR, f'Coef_min_{coef_min}_Sim_{data_id + 1}_Original_Trace.csv' ) if save_trace else None
                multi_ = SMR.Multicollinear( norm_threshold = norm_threshold, alpha = alpha,
                                                total_time_limit = total_time_limit, per_solve_time_limit = per_solve_time_limit,
                                                log_file = log_path, log_tag = log_tag, trace = [], trace_file = trace_path )
                start = time.time()
                z_pos = multi_.Bertsimas_Detection( X = X, col_names = feature_col )
                end = time.time()
                ACC, FPR = multi_.Multicollinear_score( z_pos, original_col )
                print( f'ACC: {round( ACC, 2 )}, MFPR: {round( FPR, 2 )}, execution time: {round( end - start, 2 )}' )
                _record_performance( coef_min, data_id, 'Original', ACC, FPR, end - start )
                
                # Bertsimas runs no inspection: blank Norm / Coefficient; keep Gap / Time.
                for entry in multi_.trace:
                    detected_rows.append( {
                        'Method': 'Original', 'Status': entry['status'],
                        'Detected_Relationship': _support_to_feature_str( entry['support'], feature_col ),
                        'Norm': None, 'Coefficient': None,
                        'Gap': entry['gap'], 'Time': entry['time'],
                    } )
            
            # Persist the performance table incrementally.
            pd.DataFrame( performance_ ).to_excel( os.path.join( OUTPUT_DIR, 'Performance.xlsx' ), index = False )
            
            if save_relationship:
                _build_detected_sheet( workbook = detected_wb, sheet_name = f'Sim_{data_id + 1}', rows = detected_rows )
        
        if save_relationship:
            wb.save( os.path.join( OUTPUT_DIR, f'Coef_min_{coef_min}_Ground_Truth.xlsx' ) )
            detected_wb.save( os.path.join( OUTPUT_DIR, f'Coef_min_{coef_min}_Detected.xlsx' ) )
            print( f'Saved ground-truth and detected workbooks for coef_min={coef_min}' )
    
    # Paired significance tests on the produced Performance table.
    if run_significance:
        method_order = [ m['label'] for m in METHODS ] + ['Original']
        _run_significance_test( pd.DataFrame( performance_ ), OUTPUT_DIR, significance_baseline, method_order )
    
    print( f'Detection results written to {OUTPUT_DIR}' )
    return OUTPUT_DIR


# --------------------------------------------------------------------------- #
# 2. Dimensionality-reduction screen comparison
# --------------------------------------------------------------------------- #
def run_reduction_simulation(
    BASE_DIR: str,
    n: int, p: int, MR: list,
    noise_scale: float = 0.01,
    coef_mins = ( 0.0, ),
    simulations: int = 10,
    seed: int = 911122,
    corr_outlier_thresholds = None,
    csv_id: str = None,
):
    """Compare the eigenvector screen against the correlation screen.
    
    For each simulated dataset the eigenvector reduction runs once and the
    correlation reduction runs once per entry in ``corr_outlier_thresholds``.
    Accuracy (ACC), false-positive rate (FPR), reduced feature count and time
    are written to Performance.xlsx.
    
    Parameters mirror run_detection_simulation; ``corr_outlier_thresholds`` is
    the list of z thresholds for the correlation screen, and ``alpha`` scales
    the small-eigenvalue cutoff for the eigenvector screen.
    """
    csv_id = csv_id or time.strftime( '%Y%m%d-%H%M%S', time.localtime() )
    
    if corr_outlier_thresholds is None:
        corr_outlier_thresholds = ( 3, 4 ) if p == 1000 else ( 4, 5 )
    
    OUTPUT_DIR = os.path.join( BASE_DIR, 'Reduction', f'p_{p}_{csv_id}' )
    os.makedirs( OUTPUT_DIR, exist_ok = True )
    
    rng = np.random.default_rng( seed = seed )
    
    performance_ = { 'Coef_min': [], 'Simulation_id': [], 'Method': [], 'ACC': [], 'FPR': [], 'Reduced_p': [], 'Time': [] }
    
    def _record( coef_min, data_id, label, ACC, FPR, reduced_p, elapsed ):
        performance_['Coef_min'].append( coef_min )
        performance_['Simulation_id'].append( data_id )
        performance_['Method'].append( label )
        performance_['ACC'].append( ACC )
        performance_['FPR'].append( FPR )
        performance_['Reduced_p'].append( reduced_p )
        performance_['Time'].append( elapsed )
    
    for coef_min in coef_mins:
        data_seeds = rng.integers( low = 0, high = 10000, size = simulations )
        
        for data_id, data_seed in enumerate( data_seeds ):
            print( '=.' * 50 )
            print( f'Coef_min: {coef_min}, Data: {data_id + 1} / {simulations}' )
            
            original_col, coef_dict, X, feature_col = SMR.Simulation_Data( n, p, MR, noise_scale, data_seed, coef_min = coef_min )
            
            true_feat = []
            for key, value in original_col.items():
                print( f'{key}: {value}' )
                for feat_ in value:
                    true_feat += feat_
            print( f'Total number of features involved in multicollinearity: {len( true_feat )}' )
            
            multi_ = SMR.Multicollinear()
            
            # Eigenvector screen.
            start_ = time.time()
            _, eigen_reduction_idx = multi_.Eigvec_Dimensionality_Reduction( X = X )
            eigen_time = time.time() - start_
            eigen_ACC, eigen_FPR = multi_.Reduction_score( true_idx = true_feat, reduction_idx = eigen_reduction_idx )
            print( f'Reduced_p: {len( eigen_reduction_idx )}, Eigvec: ACC = {eigen_ACC:.2f}, FPR = {eigen_FPR:.2f}, Time = {eigen_time:.2f}' )
            _record( coef_min, data_id, 'Eigvec', eigen_ACC, eigen_FPR, len( eigen_reduction_idx ), eigen_time )
            
            # Correlation screen at each threshold.
            for outlier_threshold in corr_outlier_thresholds:
                start_ = time.time()
                _, corr_reduction_idx = multi_.Corr_Dimensionality_Reduction( X = X, outlier_threshold = outlier_threshold )
                corr_time = time.time() - start_
                corr_ACC, corr_FPR = multi_.Reduction_score( true_idx = true_feat, reduction_idx = corr_reduction_idx )
                print( f'Reduced_p: {len( corr_reduction_idx )}, Corr (z = {outlier_threshold}): ACC = {corr_ACC:.2f}, FPR = {corr_FPR:.2f}, Time = {corr_time:.2f}' )
                _record( coef_min, data_id, f'Corr (z = {outlier_threshold})', corr_ACC, corr_FPR, len( corr_reduction_idx ), corr_time )
            
            pd.DataFrame( performance_ ).to_excel( os.path.join( OUTPUT_DIR, 'Performance.xlsx' ), index = False )
    
    print( f'Reduction results written to {OUTPUT_DIR}' )
    return OUTPUT_DIR


# --------------------------------------------------------------------------- #
# 3. Real-world detection (pre-processed OpenML / UCI datasets)
# --------------------------------------------------------------------------- #
def _discover_processed_datasets( data_dir: str ) -> list:
    """Return the base names of every ``*_Processed.csv`` file in ``data_dir``."""
    suffix = '_Processed.csv'
    return sorted( f[:-len( suffix )] for f in os.listdir( data_dir ) if f.endswith( suffix ) )


def run_realworld_detection(
    BASE_DIR: str,
    data_dir: str,
    datasets = None,
    run_SMR: bool = True,
    run_bertsimas: bool = True,
    reduction: bool = False,
    norm_threshold: float = 10 ** ( -2 ),
    alpha: float = 0.5,
    total_time_limit: int = None,
    per_solve_time_limit: int = None,
    csv_id: str = None,
):
    """Run detection on pre-processed real-world datasets.
    
    Each dataset is read from ``{data_dir}/{name}_Processed.csv`` and used
    directly (already pre-processed; no normalization applied). For every
    dataset the selected detector(s) run, the detected relationships are traced
    to CSV, and per-dataset counts (relationships, inequality / irreducibility
    failures, time) are collected into Overview.csv.
    
    Parameters
    ----------
    BASE_DIR : str
        Root directory for all outputs (created if missing).
    data_dir : str
        Folder containing the ``*_Processed.csv`` files.
    datasets : list of str or None
        Dataset base names (without the ``_Processed.csv`` suffix). None runs
        every processed dataset found in ``data_dir``.
    run_SMR, run_bertsimas : bool
        Enable SMR_Detection and/or Bertsimas_Detection.
    reduction : bool
        Enable the dimensionality-reduction screen inside Enhanced_Detection.
    norm_threshold, alpha : float
        Inequality norm threshold and eigenvalue-cutoff scale.
    total_time_limit, per_solve_time_limit : int or None
        Overall and per-solve Gurobi time budgets (seconds).
    csv_id : str or None
        Unused for the folder name here (kept for signature symmetry).
    """
    OUTPUT_DIR = os.path.join( BASE_DIR, 'RealWorld' )
    os.makedirs( OUTPUT_DIR, exist_ok = True )
    
    if datasets is None:
        datasets = _discover_processed_datasets( data_dir )
    
    record_df = {
        'Dataset': [], 'Method': [], 'Number of detected relationships': [],
        'Fail inequality inspection': [], 'Fail irreducibility inspection': [], 'Execution_Time': [],
    }
    
    def _detect( X, feature_col, detector, trace_path ):
        """Run one detector, trace it, and return (z_pos, ineq_fail, irre_fail, time)."""
        multi_ = SMR.Multicollinear( reduction = reduction, norm_threshold = norm_threshold, alpha = alpha,
                                        total_time_limit = total_time_limit, per_solve_time_limit = per_solve_time_limit,
                                        trace = [], trace_file = trace_path )
        start_ = time.time()
        if detector == 'SMR_Detection':
            z_pos = multi_.SMR_Main( X = X, col_names = feature_col )
        else:
            z_pos = multi_.Bertsimas_Detection( X = X, col_names = feature_col )
        elapsed = time.time() - start_
        
        ineq_fail = irre_fail = 0
        for z in z_pos:
            norm, _ = multi_._inequality_inspection( X = X, feature_idx = z )
            if norm > multi_.norm_threshold:
                ineq_fail += 1
            if not multi_._irreducibility_inspection( X = X, feature_idx = z ):
                irre_fail += 1
        return z_pos, ineq_fail, irre_fail, elapsed
    
    for dName in datasets:
        print( '=.' * 50 )
        print( f'Dataset: {dName}' )
        
        data = pd.read_csv( os.path.join( data_dir, f'{dName}_Processed.csv' ) )
        feature_col = np.array( data.columns.to_list() )
        X = data.values
        
        detectors = []
        if run_SMR:  detectors.append( ( 'SMR_Detection', f'{dName}_Detected.csv' ) )
        if run_bertsimas: detectors.append( ( 'Bertsimas_Detection', f'{dName}_Detected_Bertsimas.csv' ) )
        
        for detector, trace_name in detectors:
            print( '{:-^100}'.format( detector ) )
            z_pos, ineq_fail, irre_fail, elapsed = _detect(
                X, feature_col, detector, os.path.join( OUTPUT_DIR, trace_name ) )
            record_df['Dataset'].append( dName )
            record_df['Method'].append( detector )
            record_df['Number of detected relationships'].append( len( z_pos ) )
            record_df['Fail inequality inspection'].append( ineq_fail )
            record_df['Fail irreducibility inspection'].append( irre_fail )
            record_df['Execution_Time'].append( elapsed )
            
            # Persist Overview incrementally so long runs are not lost.
            pd.DataFrame( record_df ).to_csv( os.path.join( OUTPUT_DIR, 'Overview.csv' ), index = False )
    
    print( f'Real-world results written to {OUTPUT_DIR}' )
    return OUTPUT_DIR
